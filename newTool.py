import os
import shutil
import sys
import zipfile
from configparser import ConfigParser
import re
import xml.dom.minidom as xmldom
import cv2
import numpy as np
from skimage.metrics import structural_similarity
from PIL import Image, ImageDraw

import openpyxl
from PyQt5.QtCore import QDateTime, Qt, QTimer, QCoreApplication
from PyQt5.QtWidgets import QApplication, QGroupBox, QHBoxLayout, QLabel, QLineEdit, QListWidget, QProgressBar, \
    QPushButton, QTableWidget, QVBoxLayout, QWidget, QFileDialog, QMessageBox, QTableWidgetItem

from datetime import datetime

from openpyxl import load_workbook
from PIL import Image, ImageGrab


def get_program_path():
    print('path start...')
    return os.path.dirname(os.path.abspath(sys.argv[0]))


def get_config_file_path():
    print('config start...')
    return os.path.join(get_program_path(), ".excel_config.ini")


def load_file_paths(self):
    print('load start...')
    self.config = ConfigParser()
    self.config.read(get_config_file_path())
    return self.config['Paths'] if 'Paths' in self.config else {}


def save_file_paths(self):
    if self.parent.input_browse.text():
        config = ConfigParser()
        config['Paths'] = {'default_path': os.path.dirname(os.path.abspath(self.parent.input_browse.text()))}
        with open(get_config_file_path(), 'w') as configfile:
            config.write(configfile)


def set_error_message(content):
    QMessageBox.critical(QWidget, 'エラー', content)
    return False


def compare_images(image1_path, image2_path):
    green_color = (0, 255, 0)
    yellow_color = (0, 255, 255)
    cyan_color = (255, 255, 0)
    black_color = (0, 0, 0)

    rect_color = yellow_color
    rect_size = 2
    txt_color = black_color
    txt_size = 2

    contour_color1 = green_color
    contour_color2 = cyan_color

    offset = -15

    # Loading images using OpencV
    img1 = cv2.imread(image1_path)
    img2 = cv2.imread(image2_path)

    print("Comparing images:", image1_path, "and", image2_path)

    # 获取图片尺寸
    width1 = img1.shape[0]
    height1 = img1.shape[1]
    width2 = img2.shape[0]
    height2 = img2.shape[1]

    # 如果尺寸不同，则图片肯定不相同
    if width1 != width2 or height1 != height2:
        print("ピクセルが一致しません。")
        set_error_message("ピクセルが一致しません。")

    # Switch to gray
    img1_gray = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY)
    img2_gray = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY)

    # Computation of the structural similarity index
    (score, diff) = structural_similarity(img1_gray, img2_gray, full=True)

    print("\n\033[1;31;34mImage Similarity: {:.5f}%".format(score * 100))
    print("Image Difference: {:.5f}%".format(100 - score * 100))
    print('\033[0m')

    diff = (diff * 255).astype("uint8")
    diff_boxes = cv2.merge([diff, diff, diff])
    thresh = cv2.threshold(diff, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]

    # Contours
    contours = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    contours = contours[0] if len(contours) == 2 else contours[1]

    nb_differences = 0

    for c in contours:
        area = cv2.contourArea(c)

        if area > 50:
            txt = str(nb_differences + 1)
            x, y, w, h = cv2.boundingRect(c)

            # Adding rectangles, text and contours
            # cv2.rectangle(img1, (x, y), (x + w, y + h), rect_color, rect_size)

            cv2.rectangle(img2, (x, y), (x + w, y + h), rect_color, rect_size)
            cv2.putText(img2, txt, (x, y + h + offset), cv2.FONT_HERSHEY_SCRIPT_SIMPLEX,
                        1, txt_color, txt_size)

            # Saving images
            basename = os.path.basename(image1_path).split('.')[0]
            # image1 = "E:" + "/" + basename + "_image1.jpg"
            image2 = "E:" + "/" + basename + "_image2.jpg"

            # cv2.imwrite(image1, img1)
            cv2.imwrite(image2, img2)

            nb_differences += 1

    print("\033[1;31;91m==> Number of differences =", nb_differences, '\033[0m')

    # # 打开两张图片
    # image1 = Image.open(image1_path)
    # image2 = Image.open(image2_path)
    #
    # # 获取图片尺寸
    # width1, height1 = image1.size
    # width2, height2 = image2.size
    #
    # # 如果尺寸不同，则图片肯定不相同
    # if width1 != width2 or height1 != height2:
    #     return False
    #
    # # 逐像素比较
    # for y in range(height1):
    #     for x in range(width1):
    #         pixel1 = image1.getpixel((x, y))
    #         pixel2 = image2.getpixel((x, y))
    #         if pixel1 != pixel2:
    #             return False
    #
    # # 如果所有像素都相同，则图片相同
    # return True


class EventHandler:
    def __init__(self, parent):
        self.default_path = None
        self.config = None
        self.current_value = None
        self.current_row = None
        self.result_message = None
        self.timer_for_list = None
        self.pending_files_iterator = None
        self.pending_files = None
        self.folder_after = None
        self.parent = parent
        self.folder_before = None
        self.detail_table_timer = QTimer(self.parent)
        self.pic_no_cell = []

    def browse_button_click(self):
        print('browse begin...')
        self.default_path = load_file_paths(self)
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        folder_before, _ = QFileDialog.getOpenFileName(self.parent, "Select Excel File",
                                                       self.default_path.get('default_path', ''),
                                                       "Excel Files (*.xlsx *.xls)",
                                                       options=options)
        if folder_before:
            self.folder_before = folder_before
            self.parent.input_browse.setText(self.folder_before)
            self.parent.exec_button.setDisabled(False)

    def execute(self):
        wb = openpyxl.load_workbook(self.parent.input_browse.text())
        sheet = wb.active
        no_cells = []
        # 遍历A列中的所有单元格
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1), start=1):
            for col_idx, cell in enumerate(row, start=1):
                # 检查单元格的值是否以"No."开头
                if cell.value and str(cell.value).startswith("No."):
                    # 将值和坐标添加到列表中
                    no_cells.append({
                        'value': cell.value,
                        'row': row_idx,
                        'column': col_idx
                    })

        # 打印统计结果和单元格坐标
        for cell_info in no_cells:
            # print(f"值：{cell_info['value']}, 行：{cell_info['row']}, 列：{cell_info['column']}")
            print('no_cells', cell_info)

        wb.close()

        file_name = os.path.basename(self.parent.input_browse.text())
        new_name = str(file_name.split('.')[0]) + '.zip'
        dir_path = os.path.dirname(os.path.abspath(self.parent.input_browse.text()))
        new_path = os.path.join(dir_path, new_name)
        if os.path.exists(new_path):
            os.remove(new_path)
        shutil.copyfile(self.parent.input_browse.text(), new_path)

        file_zip = zipfile.ZipFile(new_path, 'r')
        zip_file_name = new_name.split('.')[0]
        zip_dir = os.path.join(dir_path, zip_file_name)
        for files in file_zip.namelist():
            file_zip.extract(files, zip_dir)
        file_zip.close()

        # 得到文档对象
        image_info = dict()
        dom_obj = xmldom.parse(zip_dir + os.sep + 'xl' + os.sep + 'drawings' + os.sep + 'drawing1.xml')
        # 得到元素对象
        element = dom_obj.documentElement

        def _f(subElementObj):
            self.pic_no_cell = []
            for anchor in subElementObj:
                xdr_from = anchor.getElementsByTagName('xdr:from')[0]
                pic_col = xdr_from.childNodes[0].firstChild.data  # 获取标签间的数据
                pic_row = xdr_from.childNodes[2].firstChild.data

                if anchor.getElementsByTagName('xdr:pic'):
                    embed = \
                        anchor.getElementsByTagName('xdr:pic')[0].getElementsByTagName('xdr:blipFill')[
                            0].getElementsByTagName(
                            'a:blip')[0].getAttribute('r:embed')  # 获取属性

                    self.pic_no_cell.append({
                        'value': embed,
                        'row': pic_row,
                        'column': pic_col
                    })
            for pic_no_cell_info in self.pic_no_cell:
                print('pic_no_cell', pic_no_cell_info)
                # image_info[(int(row), int(col))] = img_dict.get(int(embed.replace('rId', '')), {}).get('img_path')

        sub_twoCellAnchor = element.getElementsByTagName("xdr:twoCellAnchor")
        # sub_oneCellAnchor = element.getElementsByTagName("xdr:oneCellAnchor")
        _f(sub_twoCellAnchor)
        # _f(sub_oneCellAnchor)

        img_dict = dict()
        pic_dir = 'xl' + os.sep + 'media'  # excel变成压缩包后，再解压，图片在media目录
        pic_path = os.path.join(zip_dir, pic_dir)

        file_list = os.listdir(pic_path)
        for file in file_list:
            filepath = os.path.join(pic_path, file)
            img_index = int(re.findall(r'image(\d+)\.', filepath)[0])
            img_dict[img_index] = dict(img_index=img_index, img_path=filepath)
        print('..…..…..…..…..….', img_dict)

        for cell_info in no_cells:
            compare_list = []
            for pic_no_cell_info in self.pic_no_cell:
                print('ZZZZZZ', pic_no_cell_info)
                if cell_info['row'] == int(pic_no_cell_info['row']) + 1:
                    compare_list.append(pic_no_cell_info['value'])
            if len(compare_list) != 2:
                # ERROR
                print('compare_list_ERROR', compare_list)
            else:
                print('compare_list_COMAPRE', compare_list)
                image1_path = ''
                image2_path = ''
                for i in range(2):
                    print('pageno', i)
                    for key, value in img_dict.items():
                        if value['img_index'] == int(compare_list[i][3:]):
                            if i == 0:
                                image1_path = img_dict[int(compare_list[i][3:])]['img_path']
                            if i == 1:
                                image2_path = img_dict[int(compare_list[i][3:])]['img_path']
                            print('img_dict', img_dict[int(compare_list[i][3:])])

                check_flag = compare_images(image1_path, image2_path)

    def app_exit(self):
        save_file_paths(self)
        self.parent.close()


class MyApp(QWidget):
    def __init__(self):
        super().__init__()

        self.button_browse = None
        self.input_browse = None
        self.status_label = None
        self.tips_layout_2 = None
        self.tips_layout_1 = None
        self.current_datetime = None
        self.timer = None
        self.main_layout = None
        self.progress_bar = None
        self.tips_label = None
        self.tips_layout = None
        self.tips_group = None
        self.exit_button = None
        self.save_button = None
        self.exec_button = None
        self.button_layout = None
        self.button_group = None
        self.bottom_layout = None
        self.bottom_right_layout = None
        self.table_widget = None
        self.bottom_right_group = None
        self.list_widget = None
        self.bottom_left_layout = None
        self.bottom_left_group = None
        self.button_after = None
        self.input_field_after = None
        self.label_after = None
        self.button_before = None
        self.label_before = None
        self.top_layout = None
        self.top_layout_2 = None
        self.top_layout_1 = None
        self.top_group = None
        self.input_field_before = None
        self.event_handler = EventHandler(self)
        self.init_ui()

    def init_ui(self):
        self.top_group = QGroupBox("選択")
        self.top_layout_1 = QHBoxLayout()
        self.top_layout_2 = QHBoxLayout()
        self.top_layout = QVBoxLayout()
        self.label_before = QLabel('エクセル')
        self.input_browse = QLineEdit()
        self.input_browse.setReadOnly(True)
        self.button_browse = QPushButton('開く')
        self.button_browse.clicked.connect(self.event_handler.browse_button_click)

        self.top_layout_1.addWidget(self.label_before)
        self.top_layout_1.addWidget(self.input_browse)
        self.top_layout_1.addWidget(self.button_browse)
        self.top_layout.addLayout(self.top_layout_1)
        self.top_layout.addLayout(self.top_layout_2)
        self.top_group.setLayout(self.top_layout)

        self.bottom_right_group = QGroupBox('結果')
        self.bottom_right_layout = QVBoxLayout()
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(3)
        self.table_widget.setHorizontalHeaderLabels(['番号', '状態', '備考'])

        self.bottom_right_layout.addWidget(self.table_widget)
        self.bottom_right_group.setLayout(self.bottom_right_layout)

        self.bottom_layout = QHBoxLayout()
        self.bottom_layout.addWidget(self.bottom_right_group)

        self.button_group = QGroupBox("操作")
        self.button_layout = QHBoxLayout()
        self.exec_button = QPushButton('実行')
        self.exec_button.setDisabled(True)
        self.save_button = QPushButton('保存')
        self.save_button.setDisabled(True)
        self.exit_button = QPushButton('退出')
        self.exec_button.clicked.connect(self.event_handler.execute)
        # self.save_button.clicked.connect(self.event_handler.save_records)
        self.exit_button.clicked.connect(self.event_handler.app_exit)

        self.button_layout.addWidget(self.exec_button)
        self.button_layout.addWidget(self.save_button)
        self.button_layout.addWidget(self.exit_button)
        self.button_group.setLayout(self.button_layout)

        self.tips_group = QGroupBox("状態")
        self.tips_layout = QVBoxLayout()
        self.tips_layout_1 = QHBoxLayout()
        self.tips_layout_2 = QHBoxLayout()
        self.status_label = QLabel('画面初期化...')
        self.tips_label = QLabel('')
        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.tips_layout_1.addWidget(self.status_label)
        self.tips_layout_2.addWidget(self.tips_label)
        self.tips_layout_2.addWidget(self.progress_bar)
        self.tips_layout.addLayout(self.tips_layout_1)
        self.tips_layout.addLayout(self.tips_layout_2)
        self.tips_group.setLayout(self.tips_layout)

        self.main_layout = QVBoxLayout()
        self.main_layout.addWidget(self.top_group)
        self.main_layout.addLayout(self.bottom_layout)
        self.main_layout.addWidget(self.button_group)
        self.main_layout.addWidget(self.tips_group)

        self.setLayout(self.main_layout)
        self.setWindowTitle('BIP Evidence-Tool Ver.1.0 PyQt5')
        self.setGeometry(100, 100, 800, 600)
        self.timer_init()

    def timer_init(self):
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_datetime)
        self.timer.start(1000)

    def update_datetime(self):
        self.current_datetime = QDateTime.currentDateTime().toString(Qt.ISODate)
        self.tips_label.setText(f'{self.current_datetime}')


if __name__ == '__main__':
    app = QApplication(sys.argv)
    my_app = MyApp()
    my_app.show()
    sys.exit(app.exec_())
