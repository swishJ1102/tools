import openpyxl
from openpyxl.utils import column_index_from_string

file_path = 'G:\\新規 Microsoft Excel ワークシート.xlsx'
# file_path = 'G:\\output_0308.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['Sheet1']

write_count = 0
for write_row in ws.iter_rows(min_row=1, values_only=True):
    read_count = 0
    for read_row in ws.iter_rows(min_row=1, values_only=True):
        if read_row[3] is not None and write_row[0] == read_row[3] + "Impl" and write_row[1] == read_row[4]:
            print('write_row', write_row)
            print('read_row', read_row)
            # ws[str(read_count) + str('F')] = write_row[3]
            # ws[str(read_count) + str('G')] = write_row[4]
            ws.cell(row=read_count + 1, column=column_index_from_string('F'), value=write_row[3])
            ws.cell(row=read_count + 1, column=column_index_from_string('G'), value=write_row[4])
        read_count += 1
    write_count += 1

wb.save(file_path)
wb.close()
