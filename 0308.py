import json
import os
import re

import javalang
import openpyxl

PATH_NAME = 'cis'
BASE_FOLDER = 'E:\\BIP\\1基線庫\\03source\\00.受領資料\\受領ソース\\商品\\' + PATH_NAME
# BASE_FOLDER = 'E:\\BIP\\1基線庫\\03source\\00.受領資料\\受領ソース\\商品\\cis\\ECIS_API\\src\\main\\java\\jp\\co\\unisys\\enability\\cis\\api\\kj\\service\\'
# BASE_FOLDER = 'G:\\test'
OUTPUT_FILE = 'G:\\output_0308.txt'
JAVA_SOURCE_NAME = 'java_source_' + PATH_NAME + '.txt'
JAVA_CODE_PATH = ''
JAVA_IMPORT = []

OUTPUT_CLASS_NAME = ''
OUTPUT_LINE_NO = ''
OUTPUT_METHOD_NAME = ''
OUTPUT_CALL_CLASS_NAME = ''
OUTPUT_CALL_METHOD_NAME = ''

OUTPUT_TOTAL_LIST = []
OUTPUT_LIST = []

WRITE_COUNT = 0


def find_java_files(folder_path_to_find, word):
    java_files = []
    for root, dirs, files in os.walk(folder_path_to_find):
        for file_con in files:
            if file_con.endswith(word + '.java'):
                java_files.append(os.path.join(root, file_con))
    return java_files


def find_specific_line_from_txt(txt_path, word, offset, code_line_no):
    line_no = 0
    if offset == '':
        with open(txt_path, 'r', encoding="utf-8") as txt_context:
            for specific_context in txt_context:
                line_no += 1
                if specific_context.find("import") >= 0 or specific_context.find(".") >= 0 \
                        or specific_context.find("//") >= 0:
                    continue
                if specific_context.find(word) > 0:
                    if code_line_no > line_no:
                        return specific_context
                    elif code_line_no <= line_no:
                        print('code_line_no <= line_no')
    else:
        count_for_pix = 0
        with open(txt_path, 'r', encoding="utf-8") as txt_context:
            for specific_context in txt_context:
                count_for_pix += 1
                if count_for_pix == code_line_no + offset:
                    return specific_context
    return None


def base_file_to_find_in(word):
    with open(JAVA_SOURCE_NAME, 'r') as file_to_find:
        for line in file_to_find:
            line = line.replace('\n', '')
            if line == word + ".java":
                return True
    return False


def parse_method_body(body, method_name):
    global OUTPUT_LINE_NO, OUTPUT_CALL_CLASS_NAME, OUTPUT_CALL_METHOD_NAME, OUTPUT_LIST, OUTPUT_TOTAL_LIST
    global WRITE_COUNT
    WRITE_COUNT = 0
    for statement in body:
        if isinstance(statement, javalang.tree.StatementExpression) \
                and isinstance(statement.expression, javalang.tree.MethodInvocation):

            key_line = find_specific_line_from_txt(JAVA_CODE_PATH, statement.expression.qualifier,
                                                   '', statement.expression.position.line)
            if key_line is None:
                continue
            if key_line.find('new') > 0:
                # GK010101_SearchTodoServiceBean searchTodoServiceBean
                # = new GK010101_SearchTodoServiceBean();
                key_to_find_in_base = key_line.split(' ')[len(key_line.split(' ')) - 5]
            else:
                # private GK010101_SearchTodoBean gk010101SearchTodoBean;
                key_to_find_in_base = key_line.split(' ')[len(key_line.split(' ')) - 2]

            if len(JAVA_IMPORT) > 0:
                for x in range(len(JAVA_IMPORT)):
                    JAVA_IMPORT[x] = JAVA_IMPORT[x] + ";"
                java_import_str = ''
                for java_import in JAVA_IMPORT:
                    if java_import.find(key_to_find_in_base + ";") > 0:
                        java_import_str = str(java_import)

                if java_import_str.find('model.') > 0 or java_import_str.find('constants.') > 0 \
                        or java_import_str.find('entity.') > 0 or java_import_str.find('mapper.') > 0 \
                        or java_import_str.find('dao.') > 0:
                    continue

            # if key_to_find_in_base.find('Bean') > 0 or key_to_find_in_base.find('Object') > 0 \
            #         or key_to_find_in_base.find('Mapper') > 0 or key_to_find_in_base.find('Example') > 0 \
            #         or key_to_find_in_base.find('Dao') > 0:
            #     continue
            # if key_to_find_in_base == 'Todo' or key_to_find_in_base == 'MailMng' or key_to_find_in_base == 'Fcr' \
            #         or key_to_find_in_base == 'RmUp' or key_to_find_in_base == 'ExecuteBaseDateM' \
            #         or key_to_find_in_base == 'WorkScheduleMngMKey' or key_to_find_in_base == 'Bl' \
            #         or key_to_find_in_base == 'Deposit' or key_to_find_in_base == 'RmUpDetail' \
            #         or key_to_find_in_base == 'MlContractHist' or key_to_find_in_base == 'CalculatingUsageKey' \
            #         or key_to_find_in_base == 'Dpr' or key_to_find_in_base == 'SpmUp' \
            #         or key_to_find_in_base == 'RcUpM' or key_to_find_in_base == 'CrUp' \
            #         or key_to_find_in_base == 'FcaUpM' or key_to_find_in_base == 'Fu' \
            #         or key_to_find_in_base == 'FcrBreakdown' or key_to_find_in_base == 'CalculatingFixIn' \
            #         or key_to_find_in_base == 'CalculatingDsUsage' or key_to_find_in_base == 'CalculatingUsage' \
            #         or key_to_find_in_base == 'CalculatingFixIn' or key_to_find_in_base == 'RmUpKey' \
            #         or key_to_find_in_base == 'ContractHist' or key_to_find_in_base == 'Rqh' \
            #         or key_to_find_in_base == 'ContractAddInfo' or key_to_find_in_base == 'RqhKey' \
            #         or key_to_find_in_base == 'ClcBreakdown' or key_to_find_in_base == 'RrNgReasonMKey' \
            #         or key_to_find_in_base == 'RecUpM':
            #     continue
            if key_to_find_in_base == OUTPUT_CLASS_NAME:
                continue

            if base_file_to_find_in(key_to_find_in_base) is True:
                OUTPUT_LIST = []
                count_method_lines(key_to_find_in_base, statement.expression.member)
                print("\nMethod Call:", statement.expression.member)
                print("Method Call qualifier:", statement.expression.qualifier)
                print("Method Call LineNo.:", statement.expression.position.line)
                OUTPUT_LINE_NO = statement.expression.position.line
                OUTPUT_CALL_CLASS_NAME = key_to_find_in_base
                OUTPUT_CALL_METHOD_NAME = statement.expression.member
                OUTPUT_LIST = [OUTPUT_CLASS_NAME, OUTPUT_METHOD_NAME, OUTPUT_LINE_NO, OUTPUT_CALL_CLASS_NAME,
                               OUTPUT_CALL_METHOD_NAME]
                OUTPUT_TOTAL_LIST.append(OUTPUT_LIST)
                WRITE_COUNT += 1


def parse_imports(tree):
    global JAVA_IMPORT
    import_statements = [imp.path for imp in tree.imports]
    JAVA_IMPORT = import_statements


def parse_classes(tree, method_name):
    global OUTPUT_CLASS_NAME, OUTPUT_METHOD_NAME, OUTPUT_LINE_NO, \
        OUTPUT_CALL_CLASS_NAME, OUTPUT_CALL_METHOD_NAME, OUTPUT_LIST, OUTPUT_TOTAL_LIST, WRITE_COUNT
    for class_def in tree.types:
        print("\nClass Name:", class_def.name)
        OUTPUT_CLASS_NAME = class_def.name
        for method in class_def.methods:
            OUTPUT_LINE_NO, OUTPUT_CALL_CLASS_NAME, OUTPUT_CALL_METHOD_NAME = '', '', ''
            OUTPUT_LIST = []
            if method_name is None or (method_name is not None and method_name == method.name):
                print("\nMethod Name:", method.name)
                OUTPUT_METHOD_NAME = method.name
                if method.body:
                    parse_method_body(method.body, method_name)
            if WRITE_COUNT == 0:
                OUTPUT_LIST = [OUTPUT_CLASS_NAME, OUTPUT_METHOD_NAME]
                OUTPUT_TOTAL_LIST.append(OUTPUT_LIST)


def parse_java_file(java_code, method_name):
    try:
        tree = javalang.parse.parse(java_code)
    except Exception as e:
        print('error  :::::    ', e)
        return
    parse_imports(tree)
    parse_classes(tree, method_name)


def count_method_lines(class_name, method_name):
    count_path = find_java_files(BASE_FOLDER, class_name + "Impl")
    if len(count_path) > 0:
        with open(count_path[0], 'r', encoding="utf-8") as count_file:
            count_code = count_file.read()

        count_tree = javalang.parse.parse(count_code)
        method_lines = {}
        for _, node in count_tree.filter(javalang.tree.MethodDeclaration):
            if node.name == method_name:
                line_count = 0
                # for statement in node.body:
                #     if hasattr(statement, 'position'):
                #         line_count += statement.position.end.line - statement.position.start.line + 1
                # method_lines[method_name] = line_count
                # start_line, start_column = node.body.position[0]
                # end_line, end_column = node.body.position[1]
                # method_lines[method_name] = end_line - start_line - 1
                for path, statement in node.filter(javalang.tree.Statement):
                    line_count += statement.position[1] - statement.position[0] + 1
                method_lines[method_name] = line_count
        return method_lines


java_file = find_java_files(BASE_FOLDER, "")
OUTPUT_TOTAL_LIST = []
for file_path in java_file:
    print('java-path : [' + file_path + '] is reading...')
    JAVA_CODE_PATH = file_path
    with open(file_path, 'r', encoding="utf-8") as java_file:
        java_code = java_file.read()
        parse_java_file(java_code, None)

with open(OUTPUT_FILE, 'w') as output_file:
    for output_line in OUTPUT_TOTAL_LIST:
        output_file.write(str(output_line) + '\n')

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"
wb.save('G:\\output_0308.xlsx')

wb = openpyxl.load_workbook('G:\\output_0308.xlsx')
sht = wb.active
with open(OUTPUT_FILE, 'r') as read_file:
    write_line = read_file.read().split('\n')
    for i in range(len(write_line)):
        print('line : ', str(write_line[i]))
        if len(write_line[i]) == 0:
            break
        write_line[i] = write_line[i].replace('\'', '\"')
        res = json.loads(write_line[i])
        for j in range(len(res)):
            item = res[j]
            sht.cell(row=i + 1, column=j + 1, value=item)

wb.save('G:\\output_0308.xlsx')
wb.close()

print('done ... ')
