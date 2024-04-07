"""成果物作成"""
import datetime
import os
import shutil
import subprocess
import sys
from configparser import ConfigParser

import openpyxl
from PyQt5.QtCore import QDateTime, Qt, QTimer, QRect
from PyQt5.QtGui import QPen, QColor, QBrush, QFont, QPainter
from PyQt5.QtWidgets import QApplication, QGroupBox, QHBoxLayout, QLabel, QLineEdit, QProgressBar, \
    QPushButton, QVBoxLayout, QWidget, QFileDialog, QMessageBox, QMainWindow, QAction, QCheckBox, QRadioButton, QSlider
from openpyxl.styles import Alignment

# [Paths]
# input_browse = E:/BIP/3開発庫/00移行設計/03移行調査成果物/ドキュメントサンプル
# input_file = G:/20240406
#
# [Inputs]
# input_kinoid = CIS0001
# input_kinoname = テスト機能
# input_tantousya = 王永盛
# input_system = CIS


FOLDER_LOG = "LOG"
FOLDER_COVERAGE = "カバレッジ"
FOLDER_COMPARE = "現新比較"
FOLDER_COMPARE_OLD = "現"
FOLDER_COMPARE_NEW = "新"
FOLDER_COMPARE_FILE = "ファイル"
FOLDER_COMPARE_DB = "データ"

EXCEL_EVIDENCE = "エビデンス"
EXCEL_TEST = "単体テスト仕様書"
EXCEL_COMPARE = "手修正確認"
EXCEL_COVERAGE = "カバレッジ"
EXCEL_LIST = "成果物一覧"
EXCEL_ALL = [
    EXCEL_EVIDENCE,
    EXCEL_TEST,
    EXCEL_COMPARE,
    EXCEL_COVERAGE,
    EXCEL_LIST
]
EXCEL_DHC = "DHC"
EXCEL_LISTS = [
    '2_(機能ID_機能名)単体テストエビデンス.xlsx',
    '2_(機能ID_機能名)単体テスト仕様書.xlsx',
    '手修正確認.xlsx',
    'カバレッジ結果に関する補足説明.xlsx',
    '成果物一覧.xlsx'
]
EXCEL_LISTS_DES = []

INPUT_BROWSE = ''
INPUT_FILE = ''
INPUT_KINOID = ''
INPUT_KINONAME = ''
INPUT_TANTOUSYA = ''
INPUT_SYSTEM = ''

# todo
INPUT_BROWSE_FLAG = ''

# button stylesheet
BUTTON_STYLESHEET = 'QPushButton {background-color:rgba(255,178,0,100%);\
                                                    color: white; \
                                                    border-radius: 10px; \
                                                    border: 2px groove gray; \
                                                    border-style: outset;}\
                    QPushButton:hover{background-color:white;\
                                        color: black;}\
                                        QPushButton:pressed{background-color:rgb(85, 170, 255); \
                                        border-style: inset; }'


def get_program_path():
    """アプリのパスを取得"""
    return os.path.dirname(os.path.abspath(sys.argv[0]))


def get_config_file_path():
    """コンフィグのパスを取得"""
    return os.path.join(get_program_path(), ".eTool_config.ini")


def load_config_content(tag):
    """コンフィグをロード"""
    config = ConfigParser()
    config.read(get_config_file_path(), encoding='utf-8')
    return config[tag] if tag in config else {}


def init_config_content():
    global INPUT_BROWSE, INPUT_FILE, INPUT_KINOID, INPUT_KINONAME, INPUT_TANTOUSYA, INPUT_SYSTEM
    paths = load_config_content('Paths')
    for path in paths:
        print("path:", path)
    if paths['input_browse'] is not None:
        INPUT_BROWSE = paths['input_browse']
    if paths['input_file'] is not None:
        INPUT_FILE = paths['input_file']
    inputs = load_config_content('Inputs')
    if inputs['input_kinoid'] is not None:
        INPUT_KINOID = inputs['input_kinoid']
    if inputs['input_kinoname'] is not None:
        INPUT_KINONAME = inputs['input_kinoname']
    if inputs['input_tantousya'] is not None:
        INPUT_TANTOUSYA = inputs['input_tantousya']
    if inputs['input_system'] is not None:
        INPUT_SYSTEM = inputs['input_system']


def save_file_paths(self):
    """コンフィグにパスインフォを保存"""
    config = ConfigParser()
    config.read(get_config_file_path(), encoding='utf-8')
    if self.parent.input_browse.text():
        config.set('Paths', 'input_browse', self.parent.input_browse.text())
    if self.parent.input_file.text():
        config.set('Paths', 'input_file', self.parent.input_file.text())
    if self.parent.input_kinoid.text():
        config.set('Inputs', 'input_kinoid', self.parent.input_kinoid.text())
    if self.parent.input_kinoname.text():
        config.set('Inputs', 'input_kinoname', self.parent.input_kinoname.text())
    if self.parent.input_tantousya.text():
        config.set('Inputs', 'input_tantousya', self.parent.input_tantousya.text())
    if self.parent.input_system.text():
        config.set('Inputs', 'input_system', self.parent.input_system.text())
    with open(get_config_file_path(), 'w', encoding='utf-8') as configfile:
        config.write(configfile)


def save_tag_context(self, tag, label):
    config = ConfigParser()
    config.read(get_config_file_path(), encoding='utf-8')
    label_text = getattr(self, label).text()
    config.set(tag, label, label_text)
    with open(get_config_file_path(), 'w', encoding='utf-8') as configfile:
        config.write(configfile)


def is_null_check(self):
    """非空チェック"""
    check_flag = False
    context = ""
    if self.parent.input_browse.text() is None or self.parent.input_browse.text() == '':
        check_flag = True
        context = context + "サンプルパス\n"
    if self.parent.input_file.text() is None or self.parent.input_file.text() == '':
        check_flag = True
        context = context + "出力パス\n"
    if self.parent.input_kinoid.text() is None or self.parent.input_kinoid.text() == '':
        check_flag = True
        context = context + "機能ID\n"
    if self.parent.input_kinoname.text() is None or self.parent.input_kinoname.text() == '':
        check_flag = True
        context = context + "機能名\n"
    if self.parent.input_tantousya.text() is None or self.parent.input_tantousya.text() == '':
        check_flag = True
        context = context + "担当者\n"
    return check_flag, context


# todo
def set_message_box(message_type, title, context):
    """メッセージを反映"""
    if message_type == 'WARNING':
        QMessageBox.warning(None, title, context)
    if message_type == 'CRITICAL':
        QMessageBox.critical(None, title, context)
    if message_type == 'INFO':
        QMessageBox.information(None, title, context)


def folder_create(self):
    """フォルダを作成"""
    try:
        os.makedirs(os.path.join(self.parent.input_file.text(), self.parent.input_kinoid.text()))
    except FileExistsError:
        set_message_box("CRITICAL", "フォルダ", "フォルダがすでに存在します。")
        raise
    except OSError as e:
        set_message_box("INFO", "フォルダ", e)
        raise
    os.makedirs(os.path.join(self.parent.input_file.text(),
                             self.parent.input_kinoid.text(), FOLDER_COVERAGE))
    os.makedirs(os.path.join(self.parent.input_file.text(),
                             self.parent.input_kinoid.text(), FOLDER_LOG))
    os.makedirs(os.path.join(self.parent.input_file.text(),
                             self.parent.input_kinoid.text(), FOLDER_COMPARE))
    os.makedirs(os.path.join(self.parent.input_file.text(),
                             self.parent.input_kinoid.text(), FOLDER_COMPARE, FOLDER_COMPARE_OLD))
    os.makedirs(os.path.join(self.parent.input_file.text(), self.parent.input_kinoid.text(),
                             FOLDER_COMPARE, FOLDER_COMPARE_OLD, FOLDER_COMPARE_FILE))
    os.makedirs(os.path.join(self.parent.input_file.text(), self.parent.input_kinoid.text(),
                             FOLDER_COMPARE, FOLDER_COMPARE_OLD, FOLDER_COMPARE_DB))
    os.makedirs(os.path.join(self.parent.input_file.text(), self.parent.input_kinoid.text(),
                             FOLDER_COMPARE, FOLDER_COMPARE_NEW))
    os.makedirs(os.path.join(self.parent.input_file.text(), self.parent.input_kinoid.text(),
                             FOLDER_COMPARE, FOLDER_COMPARE_NEW, FOLDER_COMPARE_FILE))
    os.makedirs(os.path.join(self.parent.input_file.text(), self.parent.input_kinoid.text(),
                             FOLDER_COMPARE, FOLDER_COMPARE_NEW, FOLDER_COMPARE_DB))


def excel_copy(self):
    """サンプルからファイルをコピー"""
    i = 0
    file_names = os.listdir(self.parent.input_browse.text())
    for file_name in file_names:
        for module_name in EXCEL_ALL:
            if file_name.find(module_name) >= 0:
                destination_file = os.path.join(self.parent.input_file.text(),
                                                self.parent.input_kinoid.text(),
                                                os.path.basename(file_name)
                                                .replace("_機能名", "_" + self.parent.input_kinoname.text())
                                                .replace("機能ID", self.parent.input_kinoid.text()))
                shutil.copyfile(os.path.join(self.parent.input_browse.text(),
                                             file_name), destination_file)
                EXCEL_LISTS_DES.append(destination_file)
                i += 1
    if i < len(EXCEL_LISTS):
        set_message_box("INFO", "EXPLORER", "サンプルが足りないので\nチェックしてください。")


def column_letter_to_number(column_letter):
    """ディジットに変更する"""
    column_number = 0
    for char in column_letter:
        column_number = column_number * 26 + (ord(char) - ord("A") + 1)
    return column_number


def excel_format_evidence(self, path):
    wb = openpyxl.load_workbook(path)
    ws = wb['表紙']
    align = Alignment(vertical='center', horizontal='center')
    ws.cell(row=2, column=column_letter_to_number("V")).value = self.parent.input_kinoid.text()
    ws.cell(row=2, column=column_letter_to_number("V")).alignment = align
    ws.cell(row=2, column=column_letter_to_number("S")).alignment = align
    ws.cell(row=3, column=column_letter_to_number("V")).value = self.parent.input_kinoname.text()
    ws.cell(row=3, column=column_letter_to_number("V")).alignment = align
    ws.cell(row=3, column=column_letter_to_number("S")).alignment = align

    ws.cell(row=2, column=column_letter_to_number("AF")).alignment = align
    ws.cell(row=3, column=column_letter_to_number("AF")).alignment = align
    ws.cell(row=2, column=column_letter_to_number("AM")).alignment = align
    ws.cell(row=3, column=column_letter_to_number("AM")).alignment = align

    align = Alignment(vertical='center')
    ws.cell(row=2, column=column_letter_to_number("AI")).value = EXCEL_DHC + self.parent.input_tantousya.text()
    ws.cell(row=2, column=column_letter_to_number("AI")).alignment = align
    ws.cell(row=3, column=column_letter_to_number("AI")).value = EXCEL_DHC + self.parent.input_tantousya.text()
    ws.cell(row=3, column=column_letter_to_number("AI")).alignment = align
    ws.cell(row=1, column=column_letter_to_number("AP")).alignment = align

    current_date = datetime.date.today()
    formatted_date = current_date.strftime("%Y/%m/%d")
    align = Alignment(vertical='center', horizontal='right')
    ws.cell(row=2, column=column_letter_to_number("AP")).value = formatted_date
    ws.cell(row=2, column=column_letter_to_number("AP")).alignment = align
    ws.cell(row=3, column=column_letter_to_number("AP")).value = formatted_date
    ws.cell(row=3, column=column_letter_to_number("AP")).alignment = align
    wb.save(path)
    wb.close()


def excel_format(self):
    """担当者、機能ID、機能名"""
    for excel in EXCEL_LISTS_DES:
        if excel.find(EXCEL_EVIDENCE) > 0:
            excel_format_evidence(self, excel)
        if excel.find(EXCEL_TEST) > 0:
            excel_format_evidence(self, excel)
        if self.parent.progress_bar.value() >= 90:
            self.parent.progress_bar.setValue(100)
        else:
            self.parent.progress_bar.setValue(self.parent.progress_bar.value() + 10)


def svn_operate(self):
    """SVNから更新"""
    print('svn_operate start...')
    # cmd_update = 'svn update ' + self.parent.input_browse.text()
    # result = os.system(cmd_update)
    # print("svn update result : ", result)
    result = subprocess.run(['svn', '--version'], text=True, capture_output=True, check=False)
    if "svn, version" in result.stdout:
        with subprocess.Popen(['svn', 'update', self.parent.input_browse.text()],
                              stdin=subprocess.PIPE,
                              stdout=subprocess.PIPE,
                              stderr=subprocess.PIPE, text=True) as proc:
            stdout, stderr = proc.communicate()
        if proc.returncode != 0:
            set_message_box("CRITICAL", "SVN",
                            "サンプルフォルダをSVNから最新版に更新することが失敗しました、\n自分で更新してください。")
            print(f"Command '{self.parent.input_browse.text()}' "
                  f"failed with return code {proc.returncode}")
            print("Errors:", stderr)
        else:
            set_message_box("INFO", "SVN", "SVNから更新することが成功しました、\n続けてください。")
            print(f"Command '{self.parent.input_browse.text()}' executed successfully")
            print("Output:", stdout)
    else:
        set_message_box("WARNING", "SVN",
                        "コンピューターにはまだSVNコマンドラインがインストールされていません。\nインストールしてください。")


class EventHandler:
    """EventHandler"""

    def __init__(self, parent):
        self.parent = parent

    def browse_button_click(self):
        """サンプル開く"""
        try:
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            open_path = ''
            if self.parent.input_browse.text() is not None:
                open_path = self.parent.input_browse.text()
            folder_path = QFileDialog.getExistingDirectory(self.parent, "サンプルパス選択",
                                                           directory=open_path,
                                                           options=options)
            if folder_path:
                print("folder_path", folder_path)
                self.parent.input_browse.setText(folder_path)
                set_message_box("INFO", "注意", "サンプルフォルダをSVNから最新版に更新する必要があります。")
                # svn_operate(self)
        except OSError as e:
            print("An error occurred : ", e)
            raise
        except RuntimeError as e:
            print("A runtime error occurred : ", e)
            raise

    def file_button_click(self):
        """出力開く"""
        try:
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            if self.parent.input_file.text() is not None:
                open_path = self.parent.input_file.text()
            folder_path = QFileDialog.getExistingDirectory(self.parent, "出力パス選択",
                                                           directory=open_path,
                                                           options=options)
            if folder_path:
                print(folder_path)
                self.parent.input_file.setText(folder_path)
        except Exception as e:
            print("An error occurred : ", e)
            raise

    def execute(self):
        """実行"""
        self.parent.progress_bar.setValue(0)
        check_flag, context = is_null_check(self)
        if check_flag is True:
            set_message_box("WARNING", "非空チェック", context[:len(context) - 1])
            return
            # QMessageBox.warning(None, "非空チェック", context[:len(context) - 1])
        try:
            folder_create(self)
        except Exception as e:
            print("Caught an exception in some_method:", e)
            return
        self.parent.progress_bar.setValue(10)
        excel_copy(self)
        self.parent.progress_bar.setValue(20)
        excel_format(self)
        self.parent.progress_bar.setValue(100)

        self.parent.save_button.setDisabled(False)

    def records_open(self):
        """フォルダを開く"""
        os.startfile(self.parent.input_file.text())

    def app_exit(self):
        """アプリを退出"""
        msg_box = QMessageBox()
        msg_box.setWindowTitle("ツールメッセージ")
        msg_box.setText("ツールを終了したいですか。")
        msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msg_box.setDefaultButton(QMessageBox.No)
        msg_box.button(QMessageBox.Yes).setText("はい(&Y)")
        msg_box.button(QMessageBox.No).setText("いいえ(&N)")
        result = msg_box.exec_()
        if result == QMessageBox.Yes:
            save_file_paths(self)
            self.parent.close()

    def open_config_editor(self):
        # self.parent.setEnabled(False)
        self.config_editor = ConfigEditor(self.parent)
        self.config_editor.show()


class ConfigEditor(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.setWindowTitle("オプション")
        # self.config_data = config_data
        self.main_window = main_window
        self.setWindowModality(Qt.ApplicationModal)
        self.initUI()

    def initUI(self):
        self.setStyle(QApplication.style())
        layout = QVBoxLayout()
        self.label1 = QLabel("サンプルパス")
        layout.addWidget(self.label1)
        self.input1 = QLineEdit()
        layout.addWidget(self.input1)

        # 创建一个复选框
        self.cb = QCheckBox('开关按钮', self)
        self.cb.setCheckable(True)
        self.cb.setChecked(False)  # 设置默认状态为关闭
        self.cb.toggled.connect(self.onToggle)  # 连接信号和槽函数
        layout.addWidget(self.cb)

        self.label = QLabel('No option selected', self)
        layout.addWidget(self.label)

        self.radio_btn1 = QRadioButton('Option 1', self)
        self.radio_btn1.setChecked(True)  # 设置默认选中状态
        self.radio_btn1.toggled.connect(self.on_radio_button_toggled)
        layout.addWidget(self.radio_btn1)

        self.radio_btn2 = QRadioButton('Option 2', self)
        self.radio_btn2.toggled.connect(self.on_radio_button_toggled)
        layout.addWidget(self.radio_btn2)

        self.label_switch = QLabel('Slide switch is OFF', self)
        layout.addWidget(self.label_switch)

        self.slider = QSlider(Qt.Horizontal, self)
        self.slider.setMinimumWidth(200)  # 设置最小宽度
        self.slider.setRange(0, 1)  # 设置范围为 0 到 1
        self.slider.setSliderPosition(0)  # 设置初始位置为关闭状态
        self.slider.setStyleSheet("""
                    QSlider::handle:horizontal {
                        background-color: #ffffff;
                        border: 1px solid #000000;
                        width: 50px;
                        margin: -4px 0;
                        border-radius: 8px;
                    }
                    QSlider::groove:horizontal {
                        background-color: #dddddd;
                        height: 8px;
                    }
                """)
        self.slider.sliderReleased.connect(self.on_slider_released)
        self.slider.mousePressEvent = self.on_slider_clicked  # 捕获鼠标点击事件
        layout.addWidget(self.slider)

        self.switch_button = SwitchButton()
        layout.addWidget(self.switch_button)

        self.button1 = QPushButton('キャンセル')
        self.button2 = QPushButton('確認')
        layout.addWidget(self.button1)
        layout.addWidget(self.button2)
        self.setLayout(layout)
        self.setGeometry(750, 400, 300, 150)

    def onToggle(self):
        # 当复选框状态变化时调用该函数
        if self.cb.isChecked():
            print('开关打开')
        else:
            print('开关关闭')

    def on_radio_button_toggled(self):
        selected_option = None
        if self.radio_btn1.isChecked():
            selected_option = 'Option 1'
        elif self.radio_btn2.isChecked():
            selected_option = 'Option 2'

        self.label.setText(f'Selected option: {selected_option}')

    def on_slider_released(self):
        if self.slider.value() == 1:
            self.label_switch.setText('Slide switch is ON')
        else:
            self.label_switch.setText('Slide switch is OFF')

    def on_slider_clicked(self, event):
        # 单击滑动条时切换状态
        if self.slider.value() == 0:
            self.slider.setValue(1)
        else:
            self.slider.setValue(0)
        self.on_slider_released()  # 更新标签文本


class SwitchButton(QWidget):
    """自定义Switch按钮"""

    def __init__(self, parent=None):
        super(SwitchButton, self).__init__(parent)

        # 设置无边框和背景透明
        self.setWindowFlags(self.windowFlags() | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)

        self.resize(70, 30)
        self.state = False  # 按钮状态：True表示开，False表示关
        self.setFixedSize(70, 30)
        # self.clicked.connect(self.toggle_state)  # 连接点击事件

    # def toggle_state(self):
    #     self.checked = not self.checked  # 切换状态

    def mousePressEvent(self, event):
        """鼠标点击事件：用于切换按钮状态"""
        super(SwitchButton, self).mousePressEvent(event)

        self.state = False if self.state else True
        self.update()

    def paintEvent(self, event):
        """绘制按钮"""
        super(SwitchButton, self).paintEvent(event)

        # 创建绘制器并设置抗锯齿和图片流畅转换
        painter = QPainter(self)
        painter.setRenderHints(QPainter.Antialiasing | QPainter.SmoothPixmapTransform)

        # 定义字体样式
        font = QFont('Microsoft YaHei')
        font.setPixelSize(14)
        painter.setFont(font)

        # 开关为开的状态
        if self.state:
            # 绘制背景
            painter.setPen(Qt.NoPen)
            brush = QBrush(QColor('#FF475D'))
            painter.setBrush(brush)
            painter.drawRoundedRect(0, 0, self.width(), self.height(), self.height() // 2, self.height() // 2)

            # 绘制圆圈
            painter.setPen(Qt.NoPen)
            brush.setColor(QColor('#ffffff'))
            painter.setBrush(brush)
            painter.drawRoundedRect(43, 3, 24, 24, 12, 12)

            # 绘制文本
            painter.setPen(QPen(QColor('#ffffff')))
            painter.setBrush(Qt.NoBrush)
            painter.drawText(QRect(18, 4, 50, 20), Qt.AlignLeft, '开')
        # 开关为关的状态
        else:
            # 绘制背景
            painter.setPen(Qt.NoPen)
            brush = QBrush(QColor('#FFFFFF'))
            painter.setBrush(brush)
            painter.drawRoundedRect(0, 0, self.width(), self.height(), self.height() // 2, self.height() // 2)

            # 绘制圆圈
            pen = QPen(QColor('#999999'))
            pen.setWidth(1)
            painter.setPen(pen)
            painter.drawRoundedRect(3, 3, 24, 24, 12, 12)

            # 绘制文本
            painter.setBrush(Qt.NoBrush)
            painter.drawText(QRect(38, 4, 50, 20), Qt.AlignLeft, '关')


class MyApp(QMainWindow):
    """UI Class"""

    def __init__(self):
        """init"""
        super().__init__()

        self.current_datetime = None
        self.timer = QTimer()
        self.top_group = QGroupBox("選択")
        self.top_layout_1 = QHBoxLayout()
        self.top_layout_2 = QHBoxLayout()
        self.top_layout_3 = QHBoxLayout()
        self.top_layout_4 = QHBoxLayout()
        self.top_layout = QVBoxLayout()
        self.label_browse = QLabel('サンプル ')
        self.input_browse = QLineEdit()
        self.input_browse.setReadOnly(True)
        self.button_browse = QPushButton('開く')
        self.label_file = QLabel('出力パス')
        self.input_file = QLineEdit()
        self.input_file.setReadOnly(True)
        self.button_file = QPushButton('開く')
        # todo QComboBoxに変更
        self.label_kinoid = QLabel('機能ID  ')
        self.input_kinoid = QLineEdit()
        self.input_kinoid.setPlaceholderText('機能ID')
        self.label_kinoname = QLabel('機能名  ')
        self.input_kinoname = QLineEdit()
        self.input_kinoname.setPlaceholderText('機能名')
        self.label_tantousya = QLabel('担当者  ')
        self.input_tantousya = QLineEdit()
        self.input_tantousya.setPlaceholderText('担当者')
        self.label_system = QLabel('システム ')
        self.label_system.setToolTip('システム名')
        self.label_system.setToolTipDuration(2000)
        self.input_system = QLineEdit()
        self.input_system.setToolTip('ここではシステム名を記入してください。')
        self.input_system.setToolTipDuration(2000)
        self.input_system.setPlaceholderText('システム')

        self.top_layout_1.addWidget(self.label_browse)
        self.top_layout_1.addWidget(self.input_browse)
        self.top_layout_1.addWidget(self.button_browse)
        self.top_layout_2.addWidget(self.label_file)
        self.top_layout_2.addWidget(self.input_file)
        self.top_layout_2.addWidget(self.button_file)
        self.top_layout_3.addWidget(self.label_kinoid)
        self.top_layout_3.addWidget(self.input_kinoid)
        self.top_layout_3.addWidget(self.label_kinoname)
        self.top_layout_3.addWidget(self.input_kinoname)
        self.top_layout_4.addWidget(self.label_tantousya)
        self.top_layout_4.addWidget(self.input_tantousya)
        self.top_layout_4.addWidget(self.label_system)
        self.top_layout_4.addWidget(self.input_system)

        self.top_layout.addLayout(self.top_layout_1)
        self.top_layout.addLayout(self.top_layout_2)
        self.top_layout.addLayout(self.top_layout_3)
        self.top_layout.addLayout(self.top_layout_4)
        self.top_group.setLayout(self.top_layout)

        self.button_group = QGroupBox("操作")
        self.button_layout = QHBoxLayout()
        self.exec_button = QPushButton('実行')
        self.save_button = QPushButton('出力フォルダを開く')
        self.save_button.setDisabled(True)
        self.exit_button = QPushButton('退出')
        self.exit_button.setStyleSheet("background-color: lightgray")

        self.button_layout.addWidget(self.exec_button)
        self.button_layout.addWidget(self.save_button)
        self.button_layout.addWidget(self.exit_button)
        self.button_group.setLayout(self.button_layout)

        self.tips_group = QGroupBox("状態")
        self.tips_layout = QVBoxLayout()
        self.tips_layout_1 = QHBoxLayout()
        self.tips_layout_2 = QHBoxLayout()
        self.status_label = QLabel('')
        self.tips_label = QLabel('')
        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.tips_layout_1.addWidget(self.status_label)
        self.tips_layout_2.addWidget(self.tips_label)
        self.tips_layout_2.addWidget(self.progress_bar)
        # self.tips_layout.addLayout(self.tips_layout_1)
        self.tips_layout.addLayout(self.tips_layout_2)
        self.tips_group.setLayout(self.tips_layout)

        self.main_layout = QVBoxLayout()
        self.main_layout.addWidget(self.top_group)
        self.main_layout.addWidget(self.button_group)
        self.main_layout.addWidget(self.tips_group)

        menu_bar = self.menuBar()
        file_menu = menu_bar.addMenu('操作')
        self.exec_action = QAction('実行', self)
        file_menu.addAction(self.exec_action)
        self.exec_action.setEnabled(True)
        self.exec_action.setShortcut('Ctrl+E')

        self.open_action = QAction('報告を開く', self)
        file_menu.addAction(self.open_action)
        self.open_action.setEnabled(False)
        self.open_action.setShortcut('Alt+O')

        self.exit_action = QAction('退出', self)
        file_menu.addAction(self.exit_action)
        self.exit_action.setEnabled(True)
        self.exit_action.setShortcut('Alt+Q')

        config_menu = menu_bar.addMenu('設定')
        self.config_action = QAction('オプション', self)
        config_menu.addAction(self.config_action)
        self.config_action.setShortcut('Ctrl+I')

        # self.config_action.setDisabled(True)

        central_widget = QWidget()
        central_widget.setLayout(self.main_layout)
        self.setCentralWidget(central_widget)
        self.setLayout(self.main_layout)
        self.setWindowTitle('BIP-成果物作成-Ver.1.0-Powered by PyQt5')
        self.setGeometry(650, 350, 600, 300)
        self.setFixedSize(600, 300)
        self.event_handler = EventHandler(self)
        self.init_ui()

    def init_ui(self):
        """init_ui"""
        self.button_browse.clicked.connect(self.event_handler.browse_button_click)
        # self.button_browse.released.connect(self.event_handler.browse_button_released)
        self.button_file.clicked.connect(self.event_handler.file_button_click)
        self.exec_button.clicked.connect(self.event_handler.execute)
        self.save_button.clicked.connect(self.event_handler.records_open)
        self.exit_button.clicked.connect(self.event_handler.app_exit)

        self.exec_action.triggered.connect(self.event_handler.execute)
        self.open_action.triggered.connect(self.event_handler.records_open)
        self.exit_action.triggered.connect(self.event_handler.app_exit)
        self.config_action.triggered.connect(self.event_handler.open_config_editor)

        self.input_browse.textChanged.connect(self.onTextChanged)
        self.input_tantousya.focusOutEvent = self.onFocusOut

        self.timer_init()
        init_config_content()
        if INPUT_BROWSE is not None:
            self.input_browse.setText(INPUT_BROWSE)
        if INPUT_FILE is not None:
            self.input_file.setText(INPUT_FILE)
        if INPUT_KINOID is not None:
            self.input_kinoid.setText(INPUT_KINOID)
        if INPUT_KINONAME is not None:
            self.input_kinoname.setText(INPUT_KINONAME)
        if INPUT_TANTOUSYA is not None:
            self.input_tantousya.setText(INPUT_TANTOUSYA)
        if INPUT_SYSTEM is not None:
            self.input_system.setText(INPUT_SYSTEM)

    def timer_init(self):
        """タイマーコントロール"""
        self.timer.timeout.connect(self.update_datetime)
        self.timer.start(1000)

    def update_datetime(self):
        """タイマーを更新"""
        self.current_datetime = QDateTime.currentDateTime().toString(Qt.ISODate)
        self.tips_label.setText(f'{self.current_datetime}')

    def onTextChanged(self, text):
        # 处理文本变化事件
        print('文本内容变化:', text)
        save_tag_context(self, "Paths", "input_browse")

    def onFocusOut(self, event):
        # 处理光标离开事件
        text = self.input_tantousya.text()
        print('光标离开，文本为:', text)
        # 调用父类的 focusOutEvent 方法，以确保事件能够正常处理
        QLineEdit.focusOutEvent(self.input_tantousya, event)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle('Windows')  # Windows , windowsvista , Fusion
    my_app = MyApp()
    my_app.show()
    sys.exit(app.exec_())
