import datetime
import os
import re
import shutil
import sys
import xml.dom.minidom as xmldom
import zipfile
from configparser import ConfigParser

import cv2
import openpyxl
from PyQt5.QtCore import QDateTime, Qt, QTimer
from PyQt5.QtWidgets import QApplication, QGroupBox, QHBoxLayout, QLabel, QLineEdit, QProgressBar, \
    QPushButton, QTableWidget, QVBoxLayout, QWidget, QFileDialog, QMessageBox, QMainWindow, \
    QAction, QTableWidgetItem, QHeaderView
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as ImageExcel
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from skimage.metrics import structural_similarity

# from PIL import Image, ImageGrab

# [Paths]
# default_path = E:\
# default_output_path = G:
# output_compare_path = compare
#
# [Output]
# output_image = _output.png
# output_result = output.xlsx

TEMP_DIR = 'C:/temp-bip/'
TEMP_FILE = ''

input_browse = ''
report_path = ''
content_for_table = ''
col = 1
row = 1
data_status = True

# button stylesheet
button_stylesheet = 'QPushButton {background-color:rgba(255,178,0,100%);\
                                                    color: white; \
                                                    border-radius: 10px; \
                                                    border: 2px groove gray; \
                                                    border-style: outset;}\
                    QPushButton:hover{background-color:white;\
                                        color: black;}\
                                        QPushButton:pressed{background-color:rgb(85, 170, 255); \
                                        border-style: inset; }'


# Helper functions
def get_program_path():
    return os.path.dirname(os.path.abspath(sys.argv[0]))


def get_config_file_path():
    return os.path.join(get_program_path(), ".excel_config.ini")


def load_config_content(tag):
    config = ConfigParser()
    config.read(get_config_file_path())
    return config[tag] if tag in config else {}


def save_file_paths():
    global input_browse
    if os.path.exists(input_browse):
        config = ConfigParser()
        config.read(get_config_file_path())
        config.set('Paths', 'default_path', os.path.dirname(os.path.abspath(input_browse)))
        with open(get_config_file_path(), 'w') as configfile:
            config.write(configfile)


def set_error_message(content):
    QMessageBox.critical(None, "エラー", content)
    return False


def show_alert():
    app_for_alert = QApplication(sys.argv)
    alert = QMessageBox()
    alert.setWindowTitle("alert!")
    alert.setText("aaaaaaaaaaaaaaaaaaaaa")
    # alert.exec_()
    alert.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
    result = alert.exec_()
    if result == QMessageBox.Ok:
        print("Ok~~~~~~~~~~~~")
    elif result == QMessageBox.Cancel:
        print("Cancle~~~~~~~~~~~~~~~")


def tmp_files_del():
    # 目録
    if os.path.exists(input_browse.split('.')[0]):
        shutil.rmtree(input_browse.split('.')[0])
    else:
        print('目録が存在しない。')

    # ZIPファイル
    if os.path.exists(input_browse.split('.')[0] + '.zip'):
        os.remove(input_browse.split('.')[0] + '.zip')
    else:
        print('ZIPが存在しない。')


def compare_images(image1_path, image2_path):
    # logging.warning("compare_images, image1_path = %s , image2_path = %s", image1_path, image2_path)
    image2 = ''
    # 色を設定
    green_color = (0, 255, 0)
    yellow_color = (0, 255, 255)
    cyan_color = (255, 255, 0)
    black_color = (0, 0, 0)

    rect_color = yellow_color
    rect_size = 2
    txt_color = black_color
    txt_size = 2

    contour_color1 = green_color
    contour_color2 = cyan_color

    # 偏位量を設定
    offset = -15

    # イメージロード
    img1 = cv2.imread(image1_path)
    img2 = cv2.imread(image2_path)

    print("Comparing images:", image1_path, "and", image2_path)

    # イメージのサイズを取得
    width1 = img1.shape[0]
    height1 = img1.shape[1]
    width2 = img2.shape[0]
    height2 = img2.shape[1]

    # ピクセルが一致しないと、間違っていました。
    if width1 != width2 or height1 != height2:
        # show_alert()
        print("ピクセルが一致しません。")
        return None, None
        # set_error_message('ピクセルが一致しません。')

    # グレーを変換
    img1_gray = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY)
    img2_gray = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY)

    # Computation of the structural similarity index
    (score, diff) = structural_similarity(img1_gray, img2_gray, full=True)

    print("\n\033[1;31;34mイメージ 合わせること: {:.5f}%".format(score * 100))
    print("イメージ 違うこと: {:.5f}%".format(100 - score * 100))
    print('\033[0m')

    diff = (diff * 255).astype("uint8")
    diff_boxes = cv2.merge([diff, diff, diff])
    thresh = cv2.threshold(diff, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]

    # Contours
    contours = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    contours = contours[0] if len(contours) == 2 else contours[1]

    nb_differences = 0

    # イメージ保存
    basename = os.path.basename(image1_path).split('.')[0]
    # logging.warning("compare_images, basename = %s ", basename)
    # image1 = "E:" + "/" + basename + "_image1.jpg"
    global input_browse
    output_compare_dir = TEMP_FILE.split('.')[0] + "/" + load_config_content("Paths").get('output_compare_path', '')
    # logging.warning("compare_images, output_compare_dir = %s ", output_compare_dir)
    if os.path.exists(output_compare_dir):
        print('目録が存在します。')
    else:
        os.makedirs(output_compare_dir)
        global col
        global row
        col = 1
        row = 1

    for c in contours:
        area = cv2.contourArea(c)

        if area > 50:
            txt = str(nb_differences + 1)
            x, y, w, h = cv2.boundingRect(c)

            # Adding rectangles, text and contours
            # cv2.rectangle(img1, (x, y), (x + w, y + h), rect_color, rect_size)

            cv2.rectangle(img2, (x, y), (x + w, y + h), rect_color, rect_size)
            cv2.putText(img2, txt, (x, y + h + offset), cv2.FONT_HERSHEY_SCRIPT_SIMPLEX, 1, txt_color, txt_size)

            image2 = output_compare_dir + os.sep + basename + load_config_content("Output").get('output_image', '')
            # logging.warning("compare_images, image2 = %s ", image2)

            # cv2.imwrite(image1, img1)
            # 図を書き込み
            cv2.imwrite(image2, img2)

            nb_differences += 1

    print("\033[1;31;91m==> Number of differences =", nb_differences, '\033[0m')
    return image2, nb_differences


class EventHandler:
    def __init__(self, parent):
        self.evidence_file = None
        self.default_path = None
        self.config = None
        self.current_value = None
        self.current_row = None
        self.result_message = None
        self.timer_for_list = None
        self.pending_files_iterator = None
        self.pending_files = None
        self.parent = parent
        self.detail_table_timer = QTimer(self.parent)
        self.pic_no_cell = []
        # logging.basicConfig(filename='app.log', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

    def browse_button_click(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        evidence_file, _ = QFileDialog.getOpenFileName(None, "エビデンスを選択",
                                                       load_config_content("Paths").get('default_path', ''),
                                                       "Excel Files (*.xlsx *.xls)", options=options)
        if os.path.exists(evidence_file):
            self.parent.input_browse.setText(evidence_file)
            self.parent.exec_button.setDisabled(False)
            self.parent.exec_action.setEnabled(True)
            self.parent.save_button.setDisabled(True)
            self.parent.open_action.setEnabled(False)
            global input_browse
            input_browse = evidence_file
            global report_path
            report_path = evidence_file.split('.xlsx')[0] + load_config_content("Output").get('output_result', '')

    @staticmethod
    def browse_button_released():
        print("browse_button_released")

    # global input_browse
    # if os.path.exists(input_browse):
    #     save_file_paths()

    @staticmethod
    def find_no_in_excel(wb, sheet):
        no_cells = []
        # 遍历A列中的所有单元格
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1), start=1):
            for col_idx, cell in enumerate(row, start=1):
                # 检查单元格的值是否以"No."开头
                if cell.value and str(cell.value).startswith("NO."):
                    # 将值和坐标添加到列表中
                    no_cells.append({
                        'value': cell.value,
                        'row': row_idx,
                        'column': col_idx
                    })

        # 打印统计结果和单元格坐标
        for cell_info in no_cells:
            # print(f"值：{cell_info['value']}, 行：{cell_info['row']}, 列：{cell_info['column']}")
            print('no_cells', cell_info)

        wb.close()

        return no_cells

    @staticmethod
    def result_excel_ins(self, path, value):
        # output_path = input_browse.split('.')[0] + "/" + load_config_content("Output").get('output_result', '')
        try:
            if os.path.exists(report_path):
                wb = load_workbook(report_path)
                sht = wb.active
            else:
                wb = Workbook()
                sht = wb.active
            img = ImageExcel(path)
            # cnt = 81
            # img.width, img.height = (12.7 * cnt, 4.7 * cnt)
            cell = 'A' + str(int(row + 1))
            sht[cell] = value
            self.offset_img(self, img)
            sht.add_image(img)
            # logging.warning("result_excel_ins, report_path = %s", report_path)
            wb.save(report_path)
            wb.close()
            # logging.warning("result_excel_ins, close")
            os.chmod(report_path, 0o777)
        except Exception as e:
            print(f"exception '{e}' ...")

    @staticmethod
    def offset_img(self, img):
        p2e = pixels_to_EMU
        h, w = img.height, img.width
        size = XDRPositiveSize2D(p2e(w), p2e(h))
        global col, row
        marker = AnchorMarker(col=col, colOff=0, row=row, rowOff=0)
        row = row + 63
        img.anchor = OneCellAnchor(_from=marker, ext=size)

    def unzip(self):
        file_name = os.path.basename(TEMP_FILE)
        new_name = str(file_name.split('.')[0]) + '.zip'
        dir_path = os.path.dirname(os.path.abspath(TEMP_FILE))
        new_path = os.path.join(dir_path, new_name)
        # logging.warning("execute, new_path = %s", new_path)
        if os.path.exists(new_path):
            os.remove(new_path)
        shutil.copyfile(TEMP_FILE, new_path)

        file_zip = zipfile.ZipFile(new_path, 'r')
        zip_file_name = new_name.split('.')[0]
        zip_dir = os.path.join(dir_path, zip_file_name)
        for files in file_zip.namelist():
            file_zip.extract(files, zip_dir)
        file_zip.close()

        return zip_dir

    def pic_no_info_create(self, zip_dir):
        # 得到文档对象
        image_info = dict()
        dom_obj = xmldom.parse(zip_dir + os.sep + 'xl' + os.sep + 'drawings' + os.sep + 'drawing1.xml')
        # 得到元素对象
        element = dom_obj.documentElement

        # logging.warning("execute, element = %s", element)

        def _f(subElementObj):
            self.pic_no_cell = []
            for anchor in subElementObj:
                xdr_from = anchor.getElementsByTagName('xdr:from')[0]
                pic_col = xdr_from.childNodes[0].firstChild.data  # 获取标签间的数据
                pic_row = xdr_from.childNodes[2].firstChild.data

                if anchor.getElementsByTagName('xdr:pic'):
                    embed = \
                        anchor.getElementsByTagName('xdr:pic')[0].getElementsByTagName('xdr:blipFill')[
                            0].getElementsByTagName(
                            'a:blip')[0].getAttribute('r:embed')  # 获取属性
                    self.pic_no_cell.append({
                        'value': embed,
                        'row': pic_row,
                        'column': pic_col
                    })
            for pic_no_cell_info in self.pic_no_cell:
                print('pic_no_cell', pic_no_cell_info)
                # image_info[(int(row), int(col))] = img_dict.get(int(embed.replace('rId', '')), {}).get('img_path')

        sub_twoCellAnchor = element.getElementsByTagName("xdr:twoCellAnchor")
        # sub_oneCellAnchor = element.getElementsByTagName("xdr:oneCellAnchor")
        _f(sub_twoCellAnchor)
        # _f(sub_oneCellAnchor)

    def pic_compare(self, cell_info, img_dict):
        compare_list = []
        for pic_no_cell_info in self.pic_no_cell:
            print('ZZZZZZ', pic_no_cell_info)
            if cell_info['row'] == int(pic_no_cell_info['row']) + 1:
                compare_list.append(pic_no_cell_info['value'])
        if len(compare_list) != 2:
            # ERROR
            print('compare_list_ERROR', compare_list)
        else:
            print('compare_list_COMAPRE', compare_list)
            image1_path = ''
            image2_path = ''
            for i in range(2):
                print('pageno', i)
                for key, value in img_dict.items():
                    if value['img_index'] == int(compare_list[i][3:]):
                        if i == 0:
                            image1_path = img_dict[int(compare_list[i][3:])]['img_path']
                        if i == 1:
                            image2_path = img_dict[int(compare_list[i][3:])]['img_path']
                        print('img_dict', img_dict[int(compare_list[i][3:])])

            # if image1_path.find(".tmp") > 0:
            #     shutil.copyfile(image1_path, image1_path.replace(".tmp", ".png"))
            #     image1_path = image1_path.replace(".tmp", ".png")
            # if image2_path.find(".tmp") > 0:
            #     shutil.copyfile(image2_path, image2_path.replace(".tmp", ".png"))
            #     image2_path = image2_path.replace(".tmp", ".png")
            compare_image_path, nb_difference = compare_images(image1_path, image2_path)
            if compare_image_path:
                self.result_excel_ins(self, compare_image_path, cell_info['value'])
                return nb_difference
            if nb_difference == 0:
                return 0
            if compare_image_path is None and nb_difference is None:
                pass
        return None

    def execute(self):
        global input_browse, TEMP_DIR, TEMP_FILE
        if os.path.exists(input_browse):
            save_file_paths()
        """臨時フォルダーを作成"""
        if os.path.exists(TEMP_DIR) is False:
            os.makedirs(TEMP_DIR)
        TEMP_FILE = TEMP_DIR + str(datetime.datetime.now()) \
            .replace('-', '').replace('.', '').replace(' ', '').replace('-', '').replace(':', '') + ".xlsx"
        shutil.copyfile(input_browse, TEMP_FILE)
        global col, row, report_path
        col, row = 1, 1

        # self.parent.table_widget.clear()

        if not input_browse:
            set_error_message("ファイルが選択されていません。")
            return
        if not os.path.exists(input_browse):
            set_error_message("選択されたファイルが見つかりません。")
            return
        # if not os.path.exists(report_path):
        #     set_error_message("レポートの保存先が見つかりません。")
        #     return
        # logging.warning("execute, report_path = %s", report_path)
        # 実行前に、レポートを削除することが必要です。
        try:
            if os.path.exists(report_path):
                os.remove(report_path)
            else:
                print('report_pathが存在しない。')
        except Exception as e:
            QMessageBox.critical(self.parent, 'error', str(e))
        # try:
        #     if os.path.exists(report_path):
        #         os.remove(report_path)
        #     else:
        #         print('report_pathが存在しない。')
        # except OSError as e:
        #     if e.errno == errno.ENOENT:
        #         print(f"ファイル '{report_path}', ファイルない")
        #     elif e.errno == errno.EACCES:
        #         print("errno.EACCES", stat.filemode(os.stat(report_path).st_mode))
        #         print(f"ファイル '{report_path}', 権限ない")
        #     else:
        #         try:
        #             os.chmod(report_path, 0o777)
        #             os.remove(report_path)
        #             print(f"ファイル '{report_path}', 強制削除")
        #         except Exception as e:
        #             print(f"ファイル '{report_path}', 削除できません")
        #             QMessageBox.critical(self.parent, 'error', str(e))
        wb = openpyxl.load_workbook(TEMP_FILE)
        sheet = wb['現新画面比較']

        # エクセルの中で、ケース番号を探す
        no_cells = self.find_no_in_excel(wb, sheet)

        zip_dir = self.unzip()
        self.pic_no_info_create(zip_dir)

        img_dict = dict()
        pic_dir = 'xl' + os.sep + 'media'
        pic_path = os.path.join(zip_dir, pic_dir)

        file_list = os.listdir(pic_path)
        for file in file_list:
            filepath = os.path.join(pic_path, file)
            img_index = int(re.findall(r'image(\d+)\.', filepath)[0])
            img_dict[img_index] = dict(img_index=img_index, img_path=filepath)
        print('..…..…..…..…..….', img_dict)

        self.parent.table_widget.clear()
        self.parent.table_widget.setHorizontalHeaderLabels(['番号', '状態', '備考'])
        count = 1
        write_count = 0
        for cell_info in no_cells:
            nb_difference = self.pic_compare(cell_info, img_dict)
            if nb_difference is not None:
                row_count = self.parent.table_widget.rowCount()
                self.parent.table_widget.insertRow(row_count)
                self.parent.table_widget.setItem(write_count, 0, QTableWidgetItem(str(cell_info['value'])))
                if nb_difference == 0:
                    self.parent.table_widget.setItem(write_count, 1, QTableWidgetItem(str("〇")))
                    self.parent.table_widget.setItem(write_count, 2, QTableWidgetItem("よくできました。"))
                else:
                    self.parent.table_widget.setItem(write_count, 1, QTableWidgetItem(str("✕")))
                    self.parent.table_widget.setItem(write_count, 2,
                                                     QTableWidgetItem(
                                                         "キャプチャーには「" + str(nb_difference) + "」処違うことがある。"))
                self.parent.table_widget.horizontalHeader().setSectionResizeMode(
                    QHeaderView.ResizeMode.ResizeToContents)
                self.parent.table_widget.update()
                write_count += 1
            if nb_difference is None:
                row_count = self.parent.table_widget.rowCount()
                self.parent.table_widget.insertRow(row_count)
                self.parent.table_widget.setItem(write_count, 0, QTableWidgetItem(str(cell_info['value'])))
                self.parent.table_widget.setItem(write_count, 1, QTableWidgetItem(str("✕")))
                self.parent.table_widget.setItem(write_count, 2,
                                                 QTableWidgetItem("ピクセルが一致しません。"))
                self.parent.table_widget.horizontalHeader().setSectionResizeMode(
                    QHeaderView.ResizeMode.ResizeToContents)
                self.parent.table_widget.update()
                write_count += 1
            self.parent.progress_bar.setValue(int((count / len(no_cells)) * 100))
            count += 1
        # tmp_files_del()
        # self.parent.save_button.setDisabled(False)
        result = QMessageBox.information(self.parent, '完成', "報告生成")
        if result == QMessageBox.Ok:
            self.parent.save_button.setDisabled(False)
            self.parent.open_action.setEnabled(True)
            self.parent.exec_button.setDisabled(True)
            self.parent.exec_action.setEnabled(False)
        try:
            shutil.rmtree(TEMP_DIR)
        except OSError as e:
            print(f"ファイル「 {e.filename} 」の削除中にエラー「 {e.strerror} 」が発生しました。")
        self.parent.status_label.setText("実行完了...")
        self.parent.update()

    def records_open(self):
        if report_path:
            try:
                os.startfile(report_path)
            except Exception as e:
                QMessageBox.critical(self.parent, 'error', str(e))

    def app_exit(self):
        msg_box = QMessageBox()
        msg_box.setWindowTitle("ツールメッセージ")
        msg_box.setText("ツールを終了したいですか。")
        msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msg_box.setDefaultButton(QMessageBox.No)
        msg_box.button(QMessageBox.Yes).setText("はい(&Y)")
        msg_box.button(QMessageBox.No).setText("いいえ(&N)")
        result = msg_box.exec_()
        # result = QMessageBox.question(self.parent, 'exit', 'do you really wanna exit?',
        #                               QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        # result_button = QMessageBox.Yes if result == QMessageBox.Yes else QMessageBox.No
        # result_button_text = "はい" if result == QMessageBox.Yes else "いいえ"
        # result_button.setText(result_button_text)
        if result == QMessageBox.Yes:
            save_file_paths()
            self.parent.close()


class MyApp(QMainWindow):
    def __init__(self):
        super().__init__()

        self.info_action = None
        self.exit_action = None
        self.open_action = None
        self.exec_action = None
        self.button_file = None
        self.input_file = None
        self.label_file = None
        self.label_browse = None
        self.button_browse = None
        self.input_browse = None
        self.status_label = None
        self.tips_layout_2 = None
        self.tips_layout_1 = None
        self.current_datetime = None
        self.timer = None
        self.main_layout = None
        self.progress_bar = None
        self.tips_label = None
        self.tips_layout = None
        self.tips_group = None
        self.exit_button = None
        self.save_button = None
        self.exec_button = None
        self.button_layout = None
        self.button_group = None
        self.bottom_layout = None
        self.bottom_right_layout = None
        self.table_widget = None
        self.bottom_right_group = None
        self.list_widget = None
        self.bottom_left_layout = None
        self.bottom_left_group = None
        self.button_after = None
        self.input_field_after = None
        self.label_after = None
        self.button_before = None
        self.label_before = None
        self.top_layout = None
        self.top_layout_2 = None
        self.top_layout_1 = None
        self.top_group = None
        self.input_field_before = None
        self.event_handler = EventHandler(self)
        self.init_ui()

    def init_ui(self):
        global button_stylesheet
        self.top_group = QGroupBox("選択")
        self.top_layout_1 = QHBoxLayout()
        self.top_layout_2 = QHBoxLayout()
        self.top_layout = QVBoxLayout()
        self.label_browse = QLabel('エクセル')
        self.input_browse = QLineEdit()
        self.input_browse.setReadOnly(True)
        self.button_browse = QPushButton('開く')
        self.button_browse.clicked.connect(self.event_handler.browse_button_click)
        self.button_browse.released.connect(self.event_handler.browse_button_released)
        self.label_file = QLabel('仕様書')
        self.input_file = QLineEdit()
        self.input_file.setReadOnly(True)
        self.button_file = QPushButton('開く')
        self.button_file.setDisabled(True)

        self.top_layout_1.addWidget(self.label_browse)
        self.top_layout_1.addWidget(self.input_browse)
        self.top_layout_1.addWidget(self.button_browse)
        self.top_layout_2.addWidget(self.label_file)
        self.top_layout_2.addWidget(self.input_file)
        self.top_layout_2.addWidget(self.button_file)
        self.top_layout.addLayout(self.top_layout_1)
        self.top_layout.addLayout(self.top_layout_2)
        self.top_group.setLayout(self.top_layout)

        self.bottom_right_group = QGroupBox('結果')
        self.bottom_right_layout = QVBoxLayout()
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(3)
        self.table_widget.setHorizontalHeaderLabels(['番号', '状態', '備考'])
        self.table_widget.resizeColumnsToContents()
        table_font = self.table_widget.horizontalHeader().font()
        table_font.setBold(True)
        self.table_widget.horizontalHeader().setFont(table_font)

        self.bottom_right_layout.addWidget(self.table_widget)
        self.bottom_right_group.setLayout(self.bottom_right_layout)

        self.bottom_layout = QHBoxLayout()
        self.bottom_layout.addWidget(self.bottom_right_group)

        self.button_group = QGroupBox("操作")
        self.button_layout = QHBoxLayout()
        self.exec_button = QPushButton('実行')
        self.exec_button.setDisabled(True)
        # self.exec_button.setStyleSheet(button_stylesheet)
        # self.exec_button.setStyleSheet("background-color: red")
        self.save_button = QPushButton('報告を開く')
        self.save_button.setDisabled(True)
        self.exit_button = QPushButton('退出')
        self.exit_button.setStyleSheet("background-color: gray")
        self.exec_button.clicked.connect(self.event_handler.execute)
        self.save_button.clicked.connect(self.event_handler.records_open)
        self.exit_button.clicked.connect(self.event_handler.app_exit)

        self.button_layout.addWidget(self.exec_button)
        self.button_layout.addWidget(self.save_button)
        self.button_layout.addWidget(self.exit_button)
        self.button_group.setLayout(self.button_layout)

        self.tips_group = QGroupBox("状態")
        self.tips_layout = QVBoxLayout()
        self.tips_layout_1 = QHBoxLayout()
        self.tips_layout_2 = QHBoxLayout()
        self.status_label = QLabel('画面初期化...')
        self.tips_label = QLabel('')
        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        # self.progress_bar.setStyleSheet(
        #     "QProgressBar {border: 2px solid grey; border-radius: 5px; background-color: #FFFFFF; text-align:center; font-size:20px}")
        self.tips_layout_1.addWidget(self.status_label)
        self.tips_layout_2.addWidget(self.tips_label)
        self.tips_layout_2.addWidget(self.progress_bar)
        self.tips_layout.addLayout(self.tips_layout_1)
        self.tips_layout.addLayout(self.tips_layout_2)
        self.tips_group.setLayout(self.tips_layout)

        self.main_layout = QVBoxLayout()
        self.main_layout.addWidget(self.top_group)
        self.main_layout.addLayout(self.bottom_layout)
        self.main_layout.addWidget(self.button_group)
        self.main_layout.addWidget(self.tips_group)

        central_widget = QWidget()
        central_widget.setLayout(self.main_layout)
        self.setCentralWidget(central_widget)

        self.setLayout(self.main_layout)
        self.setWindowTitle('BIP Evidence-Tool Ver.1.0 PyQt5')
        self.setGeometry(600, 300, 700, 500)
        self.setFixedSize(700, 500)
        self.timer_init()

        menu_bar = self.menuBar()
        file_menu = menu_bar.addMenu('操作')
        self.exec_action = QAction('実行', self)
        self.exec_action.triggered.connect(self.event_handler.execute)
        file_menu.addAction(self.exec_action)
        self.exec_action.setEnabled(True)
        self.exec_action.setShortcut('Ctrl+E')

        self.open_action = QAction('報告を開く', self)
        self.open_action.triggered.connect(self.event_handler.records_open)
        file_menu.addAction(self.open_action)
        self.open_action.setEnabled(False)
        self.open_action.setShortcut('Alt+O')

        self.exit_action = QAction('退出', self)
        self.exit_action.triggered.connect(self.event_handler.app_exit)
        file_menu.addAction(self.exit_action)
        self.exit_action.setEnabled(True)
        self.exit_action.setShortcut('Alt+Q')

    def timer_init(self):
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_datetime)
        self.timer.start(1000)

    def update_datetime(self):
        self.current_datetime = QDateTime.currentDateTime().toString(Qt.ISODate)
        self.tips_label.setText(f'{self.current_datetime}')


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle('Windows')  # Windows , windowsvista , Fusion
    my_app = MyApp()
    my_app.show()
    sys.exit(app.exec_())
